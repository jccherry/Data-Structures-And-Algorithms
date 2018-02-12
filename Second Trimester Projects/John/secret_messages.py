import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import hashlib
from Crypto import Random
from Crypto.Cipher import AES
import base64

#AES Encryption functions -----------------------------------------
#https://gist.github.com/LoyVanBeek/5264046

BLOCK_SIZE = 32

PADDING = ' '

def pad(s):
    return s + (BLOCK_SIZE - len(s) % BLOCK_SIZE) * PADDING

def EncodeAES(plaindata, key):
    key = pad(key)
    cipher = AES.new(key)
    enc = cipher.encrypt(pad(plaindata))#cipher.encrypt(s)#
    return base64.b64encode(enc)

def DecodeAES(encodeddata, key):
    key = pad(key)
    cipher = AES.new(key)
    b64 = base64.b64decode(encodeddata)
    decrypt = cipher.decrypt(b64)
    return str(str(decrypt)[2:])[:-1]

#-------------------------------------------------------------------

#import excel data
workbook = load_workbook('user_data.xlsx')
data_worksheet = workbook.active


#figure out how many entries in excel sheet
number_of_entries = data_worksheet.max_row

#initialization of user dictionary
user_dictionary = {}



def menu():
    print("\nWelcome\n")
    print("1: Create Account\n2: Login\n3: Quit")

    while True:
        print("Enter Your Choice Below")
        user_choice = input()

        if user_choice == str(1):
            create_account()
            break
        elif user_choice == str(2):
            login()
            break
        elif user_choice == str(3):
            break

def create_account():
    username = hashlib.md5(input("Enter Username: ").encode()).hexdigest()

    password = input("Enter Password: ")
    encrypted_password = hashlib.md5(password.encode()).hexdigest()

    encrypted_message = EncodeAES(input("Enter Secret Message: "), password)

    global number_of_entries

    if not username in user_dictionary:
        number_of_entries += 1
        user_dictionary[username] = [encrypted_password,encrypted_message]
        data_worksheet['A' + str(number_of_entries)].value = username
        data_worksheet['B' + str(number_of_entries)].value = encrypted_password
        data_worksheet['C' + str(number_of_entries)].value = encrypted_message

    else:
        print("Username Taken!")
        menu()

def login():
    username_entered = ''
    while True:
        username_entered = input("Username: ")
        if hashlib.md5(username_entered.encode()).hexdigest() in user_dictionary:
            while True:
                password_entered = input("Password: ")
                #if the hash of the entered password matches the hash retrieved from the user dict
                if hashlib.md5(password_entered.encode()).hexdigest() == user_dictionary[hashlib.md5(username_entered.encode()).hexdigest()][0]:
                    print('Decrypted Message: ' + DecodeAES(user_dictionary[hashlib.md5(username_entered.encode()).hexdigest()][1],password_entered))
                    break
                else:
                    print("Incorrect Password")
            break


#RUNS AT START PAST THIS POINT--------------------------------------


#append all entries from excel sheet into user_dictionary
for i in range(1, number_of_entries):
    user_dictionary[data_worksheet['A'+str(i+1)].value] = [data_worksheet['B'+str(i+1)].value,data_worksheet['C'+str(i+1)].value]



#USER FACING CODE--------------------------------------------------
print(user_dictionary)
menu()
workbook.save('user_data.xlsx')
